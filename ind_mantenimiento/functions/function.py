import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

def export_to_excel(data, sheet_name: str):
    """_summary_

    Args:
        data (_type_): _description_
        sheet_name (str): _description_

    Returns:
        _type_: _description_
    """
    df = pd.DataFrame(data)
    excel_path = "projects\ind_mantenimiento\excel\indicadores.xlsx"
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a',if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name)
    return excel_path

def model_to_dict(model):
    """Convert SQLAlchemy model object to a dictionary."""
    return {column.name: getattr(model, column.name) for column in model.__table__.columns}

def generate_chart(excel_path,sheet_name):

    wb = load_workbook(excel_path)
    sheet = wb[sheet_name]  # Make sure this matches your sheet name

    # Assuming the columns are in the following order in your Excel sheet
    prod_per_data = Reference(sheet, min_col=2, min_row=2, max_row=sheet.max_row, max_col=2)
    qty01_data = Reference(sheet, min_col=3, min_row=2, max_row=sheet.max_row, max_col=3)
    qty02_data = Reference(sheet, min_col=4, min_row=2, max_row=sheet.max_row, max_col=4)
    resultado_data = Reference(sheet, min_col=5, min_row=2, max_row=sheet.max_row, max_col=5)
    
    # Set categories based on PROD_YEAR column
    categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)

    chart = BarChart()
    chart.add_data(prod_per_data, titles_from_data=True)
    chart.add_data(qty01_data, titles_from_data=True)
    chart.add_data(qty02_data, titles_from_data=True)
    chart.add_data(resultado_data, titles_from_data=True)
    
    chart.set_categories(categories)
    chart.title = "Grafica Indicadora"
    chart.x_axis.title = "Año_Productos"
    chart.y_axis.title = "Valores"

    # Adjust the position as needed
    sheet.add_chart(chart, "J5")
    
    wb.save(excel_path)